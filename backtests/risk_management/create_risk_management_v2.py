import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ============ STYLES ============
header_font = Font(bold=True, size=12, color="FFFFFF")
title_font = Font(bold=True, size=14, color="FFFFFF")
subtitle_font = Font(bold=True, size=12, color="FFFFFF")
normal_font = Font(size=11)
bold_font = Font(bold=True, size=11)
bold_red = Font(bold=True, size=11, color="C0392B")
bold_green = Font(bold=True, size=11, color="27AE60")

red_fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
green_fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
blue_fill = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
dark_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
orange_fill = PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")
purple_fill = PatternFill(start_color="8E44AD", end_color="8E44AD", fill_type="solid")
gold_fill = PatternFill(start_color="F1C40F", end_color="F1C40F", fill_type="solid")
silver_fill = PatternFill(start_color="BDC3C7", end_color="BDC3C7", fill_type="solid")
btc_fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
light_green = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
light_red = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
light_blue = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
light_yellow = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
light_orange = PatternFill(start_color="FDEBD0", end_color="FDEBD0", fill_type="solid")
light_purple = PatternFill(start_color="E8DAEF", end_color="E8DAEF", fill_type="solid")
light_gold = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
light_silver = PatternFill(start_color="F2F3F4", end_color="F2F3F4", fill_type="solid")
light_btc = PatternFill(start_color="FDEBD0", end_color="FDEBD0", fill_type="solid")
grey_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)

def style_header_row(ws, row, cols, fill=dark_fill):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = fill
        cell.alignment = center
        cell.border = thin_border

def style_data_cell(ws, row, col, fill=None):
    cell = ws.cell(row=row, column=col)
    cell.font = normal_font
    cell.alignment = center
    cell.border = thin_border
    if fill:
        cell.fill = fill

def add_title(ws, row, text, cols, fill=dark_fill):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1)
    cell.value = text
    cell.font = title_font
    cell.fill = fill
    cell.alignment = center
    cell.border = thin_border
    for c in range(2, cols + 1):
        ws.cell(row=row, column=c).border = thin_border
        ws.cell(row=row, column=c).fill = fill

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ========================================
# SHEET 1: Account Profile
# ========================================
ws1 = wb.active
ws1.title = "My Account Profile"
ws1.sheet_properties.tabColor = "2C3E50"

add_title(ws1, 1, "MY TRADING ACCOUNT PROFILE", 4, dark_fill)

headers = ["Parameter", "Value", "Daily Limit", "Notes"]
for c, h in enumerate(headers, 1):
    ws1.cell(row=2, column=c, value=h)
style_header_row(ws1, 2, 4)

profile = [
    ["Account Balance", "$5,000", "", "Starting capital"],
    ["Daily Loss Limit", "5%", "$250", "STOP trading if hit"],
    ["Maximum Loss Limit", "10%", "$500", "STOP + Review + Demo"],
    ["Risk Per Continuation Trade", "1.5%", "$75", "75% of capital allocation"],
    ["Risk Per Reversal Trade", "0.5%", "$25", "25% of capital allocation"],
    ["Max Trades Per Day", "3", "", "Quality over quantity"],
    ["Min Risk:Reward", "1:2", "", "Target = 2x stop distance"],
    ["Stop After Consecutive Losses", "2", "", "Prevent revenge trading"],
]

for i, row in enumerate(profile):
    r = i + 3
    for c, val in enumerate(row, 1):
        ws1.cell(row=r, column=c, value=val)
        fill = light_blue if i % 2 == 0 else None
        style_data_cell(ws1, r, c, fill)

# Daily scenarios
add_title(ws1, 12, "DAILY SCENARIOS", 4, blue_fill)
headers2 = ["Scenario", "Trades", "Result", "Account After"]
for c, h in enumerate(headers2, 1):
    ws1.cell(row=13, column=c, value=h)
style_header_row(ws1, 13, 4)

scenarios = [
    ["Worst Day (3 losses)", "L: -$75, L: -$75, L: -$75", "-$225 (4.5%)", "$4,775"],
    ["Bad Day (2 losses, 1 win T1)", "L: -$75, L: -$75, W: +$150", "$0 (breakeven)", "$5,000"],
    ["Average Day (1 loss, 2 wins T1)", "W: +$150, L: -$75, W: +$150", "+$225 (4.5%)", "$5,225"],
    ["Good Day (3 wins T1)", "W: +$150, W: +$150, W: +$150", "+$450 (9%)", "$5,450"],
    ["Best Day (3 wins T2)", "W: +$225, W: +$225, W: +$225", "+$675 (13.5%)", "$5,675"],
]

for i, row in enumerate(scenarios):
    r = 14 + i
    fills = [light_red, light_yellow, light_blue, light_green, light_green]
    for c, val in enumerate(row, 1):
        ws1.cell(row=r, column=c, value=val)
        style_data_cell(ws1, r, c, fills[i])

# Max trades calculation
add_title(ws1, 20, "HOW LIMITS PROTECT YOU", 4, red_fill)
headers3 = ["Situation", "Calculation", "Within Limit?", "Action"]
for c, h in enumerate(headers3, 1):
    ws1.cell(row=21, column=c, value=h)
style_header_row(ws1, 21, 4)

limits = [
    ["3 Continuation losses", "3 x $75 = $225", "YES (under $250)", "Close to limit - stop trading"],
    ["2 Continuation + 1 Reversal loss", "($75 x 2) + $25 = $175", "YES (under $250)", "Can take 1 more trade"],
    ["4 Continuation losses", "4 x $75 = $300", "NO (over $250!)", "Should have stopped at 3rd"],
    ["Hit $500 total loss", "10% of $5,000", "MAX LIMIT HIT", "STOP. Demo. Full review."],
]

for i, row in enumerate(limits):
    r = 22 + i
    fill = light_green if i < 2 else light_red
    for c, val in enumerate(row, 1):
        ws1.cell(row=r, column=c, value=val)
        style_data_cell(ws1, r, c, fill)

set_col_widths(ws1, [30, 30, 25, 35])

# ========================================
# SHEET 2: XAU/USD (Gold) Position Sizing
# ========================================
ws2 = wb.create_sheet("XAU-USD (Gold)")
ws2.sheet_properties.tabColor = "F1C40F"

add_title(ws2, 1, "XAU/USD (GOLD) - POSITION SIZING FOR $5,000 ACCOUNT", 6, gold_fill)

# Key facts
add_title(ws2, 3, "KEY FACTS", 6, dark_fill)
facts_h = ["Parameter", "Value", "", "", "", ""]
for c, h in enumerate(facts_h, 1):
    ws2.cell(row=4, column=c, value=h)
style_header_row(ws2, 4, 6)

gold_facts = [
    ["1 Standard Lot", "100 oz of Gold", "", "", "", ""],
    ["$1 move per 1.00 lot", "$100 profit/loss", "", "", "", ""],
    ["$1 move per 0.10 lot", "$10 profit/loss", "", "", "", ""],
    ["$1 move per 0.01 lot", "$1 profit/loss", "", "", "", ""],
    ["Typical Spread", "$0.20 - $0.50", "", "", "", ""],
    ["Typical Stop Distance", "$3 - $10", "", "", "", ""],
]

for i, row in enumerate(gold_facts):
    r = 5 + i
    for c, val in enumerate(row, 1):
        ws2.cell(row=r, column=c, value=val)
        style_data_cell(ws2, r, c, light_gold if i % 2 == 0 else None)

# Continuation table
add_title(ws2, 12, "CONTINUATION TRADES (1.5% = $75 Risk)", 6, green_fill)
headers = ["Stop Distance", "Lot Size", "If Stop Hits", "If Target 1 (1:2)", "If Target 2 (1:3)", "If Target 3 (1:5)"]
for c, h in enumerate(headers, 1):
    ws2.cell(row=13, column=c, value=h)
style_header_row(ws2, 13, 6)

gold_cont = [
    ["$2.00", "0.37 lots", "-$74", "+$148", "+$222", "+$370"],
    ["$3.00", "0.25 lots", "-$75", "+$150", "+$225", "+$375"],
    ["$5.00", "0.15 lots", "-$75", "+$150", "+$225", "+$375"],
    ["$7.00", "0.10 lots", "-$70", "+$140", "+$210", "+$350"],
    ["$10.00", "0.07 lots", "-$70", "+$140", "+$210", "+$350"],
    ["$15.00", "0.05 lots", "-$75", "+$150", "+$225", "+$375"],
]

for i, row in enumerate(gold_cont):
    r = 14 + i
    for c, val in enumerate(row, 1):
        ws2.cell(row=r, column=c, value=val)
        style_data_cell(ws2, r, c, light_green if i % 2 == 0 else None)

# Reversal table
add_title(ws2, 21, "REVERSAL TRADES (0.5% = $25 Risk)", 6, orange_fill)
headers = ["Stop Distance", "Lot Size", "If Stop Hits", "If Target 1 (1:3)", "If Target 2 (1:5)", "If Target 3 (1:8)"]
for c, h in enumerate(headers, 1):
    ws2.cell(row=22, column=c, value=h)
style_header_row(ws2, 22, 6)

gold_rev = [
    ["$2.00", "0.12 lots", "-$24", "+$72", "+$120", "+$192"],
    ["$3.00", "0.08 lots", "-$24", "+$72", "+$120", "+$192"],
    ["$5.00", "0.05 lots", "-$25", "+$75", "+$125", "+$200"],
    ["$7.00", "0.03 lots", "-$21", "+$63", "+$105", "+$168"],
    ["$10.00", "0.02 lots", "-$20", "+$60", "+$100", "+$160"],
    ["$15.00", "0.01 lots", "-$15", "+$45", "+$75", "+$120"],
]

for i, row in enumerate(gold_rev):
    r = 23 + i
    for c, val in enumerate(row, 1):
        ws2.cell(row=r, column=c, value=val)
        style_data_cell(ws2, r, c, light_orange if i % 2 == 0 else None)

# Real example
add_title(ws2, 30, "REAL TRADE EXAMPLE - GOLD SELL", 6, red_fill)
example_h = ["Step", "Detail", "Value", "", "", ""]
for c, h in enumerate(example_h, 1):
    ws2.cell(row=31, column=c, value=h)
style_header_row(ws2, 31, 6)

gold_example = [
    ["Signal", "SELL at Bearish OB", "💎 SELL"],
    ["Entry", "Sell Price", "$2,350.00"],
    ["Stop Loss", "Above OB zone", "$2,355.00"],
    ["Stop Distance", "$2,355 - $2,350", "$5.00"],
    ["Lot Size", "$75 ÷ ($5 × 100)", "0.15 lots"],
    ["Target 1 (1:2)", "$2,350 - $10", "$2,340.00 → Profit: +$150"],
    ["Target 2 (1:3)", "$2,350 - $15", "$2,335.00 → Profit: +$225"],
    ["If Stopped", "Loss = 1.5%", "-$75"],
]

for i, row in enumerate(gold_example):
    r = 32 + i
    for c, val in enumerate(row, 1):
        ws2.cell(row=r, column=c, value=val)
        fill = light_green if i >= 5 and i <= 6 else (light_red if i == 7 else (light_gold if i % 2 == 0 else None))
        style_data_cell(ws2, r, c, fill)

set_col_widths(ws2, [20, 20, 25, 25, 25, 25])

# ========================================
# SHEET 3: XAG/USD (Silver) Position Sizing
# ========================================
ws3 = wb.create_sheet("XAG-USD (Silver)")
ws3.sheet_properties.tabColor = "BDC3C7"

add_title(ws3, 1, "XAG/USD (SILVER) - POSITION SIZING FOR $5,000 ACCOUNT", 6, silver_fill)

# Key facts
add_title(ws3, 3, "KEY FACTS", 6, dark_fill)
for c, h in enumerate(facts_h, 1):
    ws3.cell(row=4, column=c, value=h)
style_header_row(ws3, 4, 6)

silver_facts = [
    ["1 Standard Lot", "5,000 oz of Silver", "", "", "", ""],
    ["$1 move per 1.00 lot", "$5,000 profit/loss", "", "", "", ""],
    ["$1 move per 0.10 lot", "$500 profit/loss", "", "", "", ""],
    ["$1 move per 0.01 lot", "$50 profit/loss", "", "", "", ""],
    ["Typical Spread", "$0.02 - $0.05", "", "", "", ""],
    ["Typical Stop Distance", "$0.20 - $1.00", "", "", "", ""],
    ["WARNING", "Silver is VERY volatile! Use smaller lots", "", "", "", ""],
]

for i, row in enumerate(silver_facts):
    r = 5 + i
    for c, val in enumerate(row, 1):
        ws3.cell(row=r, column=c, value=val)
        fill = light_red if i == 6 else (light_silver if i % 2 == 0 else None)
        style_data_cell(ws3, r, c, fill)

# Continuation table
add_title(ws3, 13, "CONTINUATION TRADES (1.5% = $75 Risk)", 6, green_fill)
headers = ["Stop Distance", "Lot Size", "If Stop Hits", "If Target 1 (1:2)", "If Target 2 (1:3)", "If Target 3 (1:5)"]
for c, h in enumerate(headers, 1):
    ws3.cell(row=14, column=c, value=h)
style_header_row(ws3, 14, 6)

silver_cont = [
    ["$0.10", "0.15 lots", "-$75", "+$150", "+$225", "+$375"],
    ["$0.20", "0.07 lots", "-$70", "+$140", "+$210", "+$350"],
    ["$0.30", "0.05 lots", "-$75", "+$150", "+$225", "+$375"],
    ["$0.50", "0.03 lots", "-$75", "+$150", "+$225", "+$375"],
    ["$0.75", "0.02 lots", "-$75", "+$150", "+$225", "+$375"],
    ["$1.00", "0.01 lots", "-$50", "+$100", "+$150", "+$250"],
]

for i, row in enumerate(silver_cont):
    r = 15 + i
    for c, val in enumerate(row, 1):
        ws3.cell(row=r, column=c, value=val)
        style_data_cell(ws3, r, c, light_green if i % 2 == 0 else None)

# Reversal table
add_title(ws3, 22, "REVERSAL TRADES (0.5% = $25 Risk)", 6, orange_fill)
for c, h in enumerate(headers, 1):
    ws3.cell(row=23, column=c, value=h)
style_header_row(ws3, 23, 6)

silver_rev = [
    ["$0.10", "0.05 lots", "-$25", "+$75", "+$125", "+$200"],
    ["$0.20", "0.02 lots", "-$20", "+$40", "+$60", "+$100"],
    ["$0.30", "0.01 lots", "-$15", "+$30", "+$45", "+$75"],
    ["$0.50", "0.01 lots", "-$25", "+$50", "+$75", "+$125"],
    ["$0.75", "0.01 lots", "-$37 (OVER!)", "⚠️ Skip", "⚠️ Skip", "⚠️ Skip"],
    ["$1.00", "0.01 lots", "-$50 (OVER!)", "⚠️ Skip", "⚠️ Skip", "⚠️ Skip"],
]

for i, row in enumerate(silver_rev):
    r = 24 + i
    for c, val in enumerate(row, 1):
        ws3.cell(row=r, column=c, value=val)
        fill = light_red if i >= 4 else (light_orange if i % 2 == 0 else None)
        style_data_cell(ws3, r, c, fill)

# Real example
add_title(ws3, 31, "REAL TRADE EXAMPLE - SILVER BUY", 6, blue_fill)
example_h2 = ["Step", "Detail", "Value", "", "", ""]
for c, h in enumerate(example_h2, 1):
    ws3.cell(row=32, column=c, value=h)
style_header_row(ws3, 32, 6)

silver_example = [
    ["Signal", "BUY at Bullish OB", "💎 BUY"],
    ["Entry", "Buy Price", "$31.00"],
    ["Stop Loss", "Below OB zone", "$30.70"],
    ["Stop Distance", "$31.00 - $30.70", "$0.30"],
    ["Pip Value", "0.01 lot = $50/dollar", "$0.30 × 5000 = $1,500/lot"],
    ["Lot Size", "$75 ÷ $1,500 per lot", "0.05 lots"],
    ["Target 1 (1:2)", "$31.00 + $0.60", "$31.60 → Profit: +$150"],
    ["Target 2 (1:3)", "$31.00 + $0.90", "$31.90 → Profit: +$225"],
    ["If Stopped", "Loss = 1.5%", "-$75"],
]

for i, row in enumerate(silver_example):
    r = 33 + i
    for c, val in enumerate(row, 1):
        ws3.cell(row=r, column=c, value=val)
        fill = light_green if i >= 6 and i <= 7 else (light_red if i == 8 else (light_silver if i % 2 == 0 else None))
        style_data_cell(ws3, r, c, fill)

set_col_widths(ws3, [20, 20, 25, 25, 25, 25])

# ========================================
# SHEET 4: BTC/USD Position Sizing
# ========================================
ws4 = wb.create_sheet("BTC-USD (Bitcoin)")
ws4.sheet_properties.tabColor = "F39C12"

add_title(ws4, 1, "BTC/USD (BITCOIN) - POSITION SIZING FOR $5,000 ACCOUNT", 6, btc_fill)

# Key facts
add_title(ws4, 3, "KEY FACTS", 6, dark_fill)
for c, h in enumerate(facts_h, 1):
    ws4.cell(row=4, column=c, value=h)
style_header_row(ws4, 4, 6)

btc_facts = [
    ["1 Lot", "1 BTC (varies by broker)", "", "", "", ""],
    ["$100 move per 1.00 lot", "$100 profit/loss", "", "", "", ""],
    ["$100 move per 0.10 lot", "$10 profit/loss", "", "", "", ""],
    ["$100 move per 0.01 lot", "$1 profit/loss", "", "", "", ""],
    ["Typical Spread", "$20 - $80", "", "", "", ""],
    ["Typical Stop Distance", "$200 - $1,500", "", "", "", ""],
    ["WARNING", "BTC is extremely volatile! Check broker lot specs", "", "", "", ""],
]

for i, row in enumerate(btc_facts):
    r = 5 + i
    for c, val in enumerate(row, 1):
        ws4.cell(row=r, column=c, value=val)
        fill = light_red if i == 6 else (light_btc if i % 2 == 0 else None)
        style_data_cell(ws4, r, c, fill)

# Continuation
add_title(ws4, 13, "CONTINUATION TRADES (1.5% = $75 Risk)", 6, green_fill)
headers = ["Stop Distance", "Lot Size", "If Stop Hits", "If Target 1 (1:2)", "If Target 2 (1:3)", "If Target 3 (1:5)"]
for c, h in enumerate(headers, 1):
    ws4.cell(row=14, column=c, value=h)
style_header_row(ws4, 14, 6)

btc_cont = [
    ["$150", "0.50 lots", "-$75", "+$150", "+$225", "+$375"],
    ["$200", "0.37 lots", "-$74", "+$148", "+$222", "+$370"],
    ["$300", "0.25 lots", "-$75", "+$150", "+$225", "+$375"],
    ["$500", "0.15 lots", "-$75", "+$150", "+$225", "+$375"],
    ["$750", "0.10 lots", "-$75", "+$150", "+$225", "+$375"],
    ["$1,000", "0.07 lots", "-$70", "+$140", "+$210", "+$350"],
    ["$1,500", "0.05 lots", "-$75", "+$150", "+$225", "+$375"],
]

for i, row in enumerate(btc_cont):
    r = 15 + i
    for c, val in enumerate(row, 1):
        ws4.cell(row=r, column=c, value=val)
        style_data_cell(ws4, r, c, light_green if i % 2 == 0 else None)

# Reversal
add_title(ws4, 23, "REVERSAL TRADES (0.5% = $25 Risk)", 6, orange_fill)
for c, h in enumerate(headers, 1):
    ws4.cell(row=24, column=c, value=h)
style_header_row(ws4, 24, 6)

btc_rev = [
    ["$150", "0.16 lots", "-$24", "+$72", "+$120", "+$192"],
    ["$200", "0.12 lots", "-$24", "+$72", "+$120", "+$192"],
    ["$300", "0.08 lots", "-$24", "+$72", "+$120", "+$192"],
    ["$500", "0.05 lots", "-$25", "+$75", "+$125", "+$200"],
    ["$750", "0.03 lots", "-$22", "+$66", "+$110", "+$176"],
    ["$1,000", "0.02 lots", "-$20", "+$60", "+$100", "+$160"],
    ["$1,500", "0.01 lots", "-$15", "+$45", "+$75", "+$120"],
]

for i, row in enumerate(btc_rev):
    r = 25 + i
    for c, val in enumerate(row, 1):
        ws4.cell(row=r, column=c, value=val)
        style_data_cell(ws4, r, c, light_orange if i % 2 == 0 else None)

# Real example
add_title(ws4, 33, "REAL TRADE EXAMPLE - BTC SELL", 6, red_fill)
for c, h in enumerate(example_h, 1):
    ws4.cell(row=34, column=c, value=h)
style_header_row(ws4, 34, 6)

btc_example = [
    ["Signal", "SELL at Bearish OB", "💎 SELL"],
    ["Entry", "Sell Price", "$87,000"],
    ["Stop Loss", "Above OB zone", "$87,500"],
    ["Stop Distance", "$87,500 - $87,000", "$500"],
    ["Lot Size", "$75 ÷ $500", "0.15 lots"],
    ["Target 1 (1:2)", "$87,000 - $1,000", "$86,000 → Profit: +$150"],
    ["Target 2 (1:3)", "$87,000 - $1,500", "$85,500 → Profit: +$225"],
    ["If Stopped", "Loss = 1.5%", "-$75"],
]

for i, row in enumerate(btc_example):
    r = 35 + i
    for c, val in enumerate(row, 1):
        ws4.cell(row=r, column=c, value=val)
        fill = light_green if i >= 5 and i <= 6 else (light_red if i == 7 else (light_btc if i % 2 == 0 else None))
        style_data_cell(ws4, r, c, fill)

set_col_widths(ws4, [20, 20, 25, 25, 25, 25])

# ========================================
# SHEET 5: Core Risk Rules
# ========================================
ws5 = wb.create_sheet("Risk Rules")
ws5.sheet_properties.tabColor = "C0392B"

add_title(ws5, 1, "CORE RISK MANAGEMENT RULES ($5,000 ACCOUNT)", 4, red_fill)

headers = ["Rule #", "Rule", "Your Limit", "Example"]
for c, h in enumerate(headers, 1):
    ws5.cell(row=2, column=c, value=h)
style_header_row(ws5, 2, 4)

rules = [
    ["1", "Max risk per Continuation trade: 1.5%", "$75", "Gold 0.15 lots with $5 stop"],
    ["2", "Max risk per Reversal trade: 0.5%", "$25", "Gold 0.05 lots with $5 stop"],
    ["3", "Max 3 positions open at once", "$225 max exposure", "3 x $75 Continuation trades"],
    ["4", "Daily loss limit: 5%", "$250", "Hit $250 loss → STOP for today"],
    ["5", "Maximum loss limit: 10%", "$500", "Hit $500 total → STOP + Demo + Review"],
    ["6", "Max 3 trades per day", "3 trades", "Even if 5 signals appear, only take 3"],
    ["7", "Stop after 2 consecutive losses", "2 losses", "L + L → Done for the day"],
    ["8", "Min Risk:Reward = 1:2", "2x target", "Stop $5 → Target $10 minimum"],
    ["9", "Move stop to breakeven at 1:1", "At +$75 profit", "Reduces risk to $0"],
    ["10", "Take 50% profit at Target 1", "At 1:2 R:R", "Lock in $75, trail the rest"],
]

for i, row in enumerate(rules):
    r = i + 3
    for c, val in enumerate(row, 1):
        ws5.cell(row=r, column=c, value=val)
        fill = light_red if i % 2 == 0 else None
        style_data_cell(ws5, r, c, fill)

set_col_widths(ws5, [12, 40, 25, 40])

# ========================================
# SHEET 6: Trade Management
# ========================================
ws6 = wb.create_sheet("Trade Management")
ws6.sheet_properties.tabColor = "27AE60"

add_title(ws6, 1, "TRADE MANAGEMENT STEPS ($5,000 ACCOUNT)", 5, green_fill)

headers = ["Step", "When", "Action", "Your Account Impact", "Example (Gold)"]
for c, h in enumerate(headers, 1):
    ws6.cell(row=2, column=c, value=h)
style_header_row(ws6, 2, 5)

mgmt = [
    ["1. Enter Trade", "Signal appears + checklist passed", "Enter at candle close", "Risk: -$75 (1.5%)", "SELL Gold at $2,350, SL $2,355"],
    ["2. Breakeven", "Price reaches 1:1", "Move stop to entry", "Risk: $0 (free trade!)", "Price hits $2,345, move SL to $2,350"],
    ["3. Take 50%", "Price reaches 1:2 (Target 1)", "Close half position", "+$75 locked in", "Close 0.075 lots at $2,340"],
    ["4. Trail Stop", "After partial profit", "Move stop with structure", "Protecting profits", "Move SL to $2,342 as price drops"],
    ["5. Final Exit", "Target 2 or trail stop hit", "Close remaining position", "+$150 to +$375 total", "Close at $2,335 = +$225 total"],
]

for i, row in enumerate(mgmt):
    r = i + 3
    for c, val in enumerate(row, 1):
        ws6.cell(row=r, column=c, value=val)
        style_data_cell(ws6, r, c, light_green if i % 2 == 0 else None)

set_col_widths(ws6, [18, 30, 25, 25, 35])

# ========================================
# SHEET 7: Drawdown Recovery
# ========================================
ws7 = wb.create_sheet("Drawdown Recovery")
ws7.sheet_properties.tabColor = "E67E22"

add_title(ws7, 1, "DRAWDOWN RECOVERY TABLE ($5,000 ACCOUNT)", 5, orange_fill)

headers = ["Loss %", "Account Balance", "Amount Lost", "Gain Needed", "Action"]
for c, h in enumerate(headers, 1):
    ws7.cell(row=2, column=c, value=h)
style_header_row(ws7, 2, 5)

drawdown = [
    ["2%", "$4,900", "$100", "2.04%", "Normal — continue trading"],
    ["5% (Daily limit)", "$4,750", "$250", "5.26%", "STOP for today, review trades"],
    ["7%", "$4,650", "$350", "7.5%", "Reduce lot size by 25%, High Only signals"],
    ["10% (MAX LIMIT)", "$4,500", "$500", "11.1%", "STOP TRADING. Demo. Full review."],
    ["15%", "$4,250", "$750", "17.6%", "Should NEVER reach here if following rules"],
    ["20%", "$4,000", "$1,000", "25%", "Serious problem — get mentor help"],
    ["50%", "$2,500", "$2,500", "100%", "Devastating — complete system failure"],
]

for i, row in enumerate(drawdown):
    r = i + 3
    if i <= 1:
        fill = light_green
    elif i <= 3:
        fill = light_orange
    else:
        fill = light_red
    for c, val in enumerate(row, 1):
        ws7.cell(row=r, column=c, value=val)
        style_data_cell(ws7, r, c, fill)

# Recovery plan
add_title(ws7, 11, "RECOVERY PLAN", 5, red_fill)
headers2 = ["Level", "First Action", "Lot Size Change", "Signal Quality", "Duration"]
for c, h in enumerate(headers2, 1):
    ws7.cell(row=12, column=c, value=h)
style_header_row(ws7, 12, 5)

recovery = [
    ["Down 5%", "Review today's trades", "Keep normal ($75)", "High Only", "Resume tomorrow"],
    ["Down 10%", "STOP immediately", "Cut to 50% ($37)", "Perfect Only", "Min 3 days break"],
    ["Down 15%", "STOP for 1 week", "Cut to 25% ($18)", "Perfect Only", "1 week demo first"],
    ["Down 20%+", "STOP for 2 weeks", "Minimum size ($10)", "Perfect Only + Mentor", "2 weeks demo"],
]

for i, row in enumerate(recovery):
    r = 13 + i
    fills = [light_green, light_yellow, light_orange, light_red]
    for c, val in enumerate(row, 1):
        ws7.cell(row=r, column=c, value=val)
        style_data_cell(ws7, r, c, fills[i])

set_col_widths(ws7, [20, 30, 25, 25, 25])

# ========================================
# SHEET 8: Signal Guide
# ========================================
ws8 = wb.create_sheet("Signal Guide")
ws8.sheet_properties.tabColor = "8E44AD"

add_title(ws8, 1, "SIGNAL TYPES & ENTRY CHECKLIST", 5, purple_fill)

headers = ["Signal", "Label", "Meaning", "Risk per Trade", "Lot Size Guide (Gold $5 stop)"]
for c, h in enumerate(headers, 1):
    ws8.cell(row=2, column=c, value=h)
style_header_row(ws8, 2, 5)

signals = [
    ["💎 BUY", "Blue (Up)", "Bullish continuation (OB/FVG)", "$75 (1.5%)", "0.15 lots"],
    ["💎 SELL", "Orange (Down)", "Bearish continuation (OB/FVG)", "$75 (1.5%)", "0.15 lots"],
    ["🚀 BUY REVERSAL", "Lime (Up)", "Bullish reversal (LQ + BOS)", "$25 (0.5%)", "0.05 lots"],
    ["🔻 SELL REVERSAL", "Fuchsia (Down)", "Bearish reversal (LQ + BOS)", "$25 (0.5%)", "0.05 lots"],
]

for i, row in enumerate(signals):
    r = i + 3
    fill = light_blue if i < 2 else light_purple
    for c, val in enumerate(row, 1):
        ws8.cell(row=r, column=c, value=val)
        style_data_cell(ws8, r, c, fill)

# Checklist
add_title(ws8, 8, "ENTRY CONFIRMATION CHECKLIST", 5, dark_fill)
headers2 = ["#", "Check Item", "For BUY", "For SELL", "✓"]
for c, h in enumerate(headers2, 1):
    ws8.cell(row=9, column=c, value=h)
style_header_row(ws8, 9, 5)

checklist = [
    ["1", "HTF Bias matches direction", "Must be 🟢 BULLISH", "Must be 🔴 BEARISH", "☐"],
    ["2", "Signal quality: High/Perfect", "High Only or Perfect Only", "High Only or Perfect Only", "☐"],
    ["3", "Clear zone visible on chart", "Green OB or Blue FVG", "Red OB or Purple FVG", "☐"],
    ["4", "Confirmation candle", "Bullish engulfing / 70%+ body", "Bearish engulfing / 70%+ body", "☐"],
    ["5", "Risk:Reward min 1:2", "Target ≥ 2x stop", "Target ≥ 2x stop", "☐"],
    ["6", "No major news in 30 min", "Check economic calendar", "Check economic calendar", "☐"],
    ["7", "Good session (London/NY)", "6:30 PM - 3:30 AM IST", "6:30 PM - 3:30 AM IST", "☐"],
    ["8", "Position size calculated", "Max $75 risk (cont)", "Max $75 risk (cont)", "☐"],
]

for i, row in enumerate(checklist):
    r = 10 + i
    for c, val in enumerate(row, 1):
        ws8.cell(row=r, column=c, value=val)
        style_data_cell(ws8, r, c, light_yellow if i % 2 == 0 else None)

set_col_widths(ws8, [8, 30, 30, 30, 8])

# ========================================
# SHEET 9: Trade Journal
# ========================================
ws9 = wb.create_sheet("Trade Journal")
ws9.sheet_properties.tabColor = "2C3E50"

add_title(ws9, 1, "TRADE JOURNAL - $5,000 ACCOUNT", 10, dark_fill)

headers = ["Date", "Pair", "Signal Type", "HTF Bias", "Entry", "Stop Loss", "Target", "Lot Size", "Result", "P&L ($)"]
for c, h in enumerate(headers, 1):
    ws9.cell(row=2, column=c, value=h)
style_header_row(ws9, 2, 10)

for i in range(30):
    r = i + 3
    for c in range(1, 11):
        cell = ws9.cell(row=r, column=c, value="")
        cell.border = thin_border
        cell.alignment = center
        if i % 2 == 0:
            cell.fill = grey_fill

# Running total section
add_title(ws9, 34, "DAILY SUMMARY", 10, dark_fill)
summary_h = ["Date", "Trades Taken", "Wins", "Losses", "Daily P&L", "Account Balance", "Daily Risk Used", "Within Limit?", "Notes", ""]
for c, h in enumerate(summary_h, 1):
    ws9.cell(row=35, column=c, value=h)
style_header_row(ws9, 35, 10)

for i in range(15):
    r = 36 + i
    for c in range(1, 11):
        cell = ws9.cell(row=r, column=c, value="")
        cell.border = thin_border
        cell.alignment = center
        if i % 2 == 0:
            cell.fill = grey_fill

set_col_widths(ws9, [12, 12, 15, 12, 12, 12, 12, 12, 12, 12])

# ========================================
# SHEET 10: Quick Reference Card
# ========================================
ws10 = wb.create_sheet("Quick Reference")
ws10.sheet_properties.tabColor = "F39C12"

add_title(ws10, 1, "QUICK REFERENCE - $5,000 ACCOUNT", 4, orange_fill)

# Cheat sheet
sections = [
    ("YOUR RISK LIMITS", red_fill, [
        ["Per Continuation Trade", "$75 (1.5%)", "Per Reversal Trade", "$25 (0.5%)"],
        ["Daily Loss Limit", "$250 (5%)", "Max Total Loss", "$500 (10%)"],
        ["Max Trades/Day", "3", "Stop After", "2 consecutive losses"],
    ]),
    ("QUICK LOT SIZE — GOLD (XAU/USD)", gold_fill, [
        ["$3 stop (Cont)", "0.25 lots", "$3 stop (Rev)", "0.08 lots"],
        ["$5 stop (Cont)", "0.15 lots", "$5 stop (Rev)", "0.05 lots"],
        ["$10 stop (Cont)", "0.07 lots", "$10 stop (Rev)", "0.02 lots"],
    ]),
    ("QUICK LOT SIZE — SILVER (XAG/USD)", silver_fill, [
        ["$0.20 stop (Cont)", "0.07 lots", "$0.20 stop (Rev)", "0.02 lots"],
        ["$0.30 stop (Cont)", "0.05 lots", "$0.30 stop (Rev)", "0.01 lots"],
        ["$0.50 stop (Cont)", "0.03 lots", "$0.50 stop (Rev)", "0.01 lots"],
    ]),
    ("QUICK LOT SIZE — BITCOIN (BTC/USD)", btc_fill, [
        ["$300 stop (Cont)", "0.25 lots", "$300 stop (Rev)", "0.08 lots"],
        ["$500 stop (Cont)", "0.15 lots", "$500 stop (Rev)", "0.05 lots"],
        ["$1000 stop (Cont)", "0.07 lots", "$1000 stop (Rev)", "0.02 lots"],
    ]),
    ("TRADE MANAGEMENT", green_fill, [
        ["At 1:1 profit", "Move stop to breakeven", "", "Risk = $0"],
        ["At 1:2 (Target 1)", "Close 50% position", "", "Lock $75 profit"],
        ["After Target 1", "Trail stop with structure", "", "Catch big moves"],
    ]),
    ("FORMULA", dark_fill, [
        ["Step 1", "Risk Amount = $5,000 × 1.5% = $75", "", ""],
        ["Step 2", "Lot Size = $75 ÷ (Stop × Pip Value)", "", ""],
        ["Step 3", "Verify: Lot × Stop × PipVal ≤ $75", "", ""],
    ]),
]

current_row = 3
for title, fill, rows in sections:
    add_title(ws10, current_row, title, 4, fill)
    current_row += 1
    for row in rows:
        for c, val in enumerate(row, 1):
            cell = ws10.cell(row=current_row, column=c, value=val)
            cell.font = normal_font
            cell.alignment = center
            cell.border = thin_border
        current_row += 1
    current_row += 1

set_col_widths(ws10, [25, 30, 25, 25])

# ========================================
# SAVE
# ========================================
filepath = r"C:\Users\Client\Desktop\ICT_Risk_Management_Guide.xlsx"
wb.save(filepath)
print(f"File saved: {filepath}")
print("Sheets: " + ", ".join([ws.title for ws in wb.worksheets]))
