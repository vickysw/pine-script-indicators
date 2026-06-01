import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ============ STYLES ============
header_font = Font(bold=True, size=12, color="FFFFFF")
title_font = Font(bold=True, size=14, color="FFFFFF")
normal_font = Font(size=11)
bold_font = Font(bold=True, size=11)

red_fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
green_fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
blue_fill = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
dark_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
orange_fill = PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")
purple_fill = PatternFill(start_color="8E44AD", end_color="8E44AD", fill_type="solid")
light_green = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
light_red = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
light_blue = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
light_yellow = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
light_orange = PatternFill(start_color="FDEBD0", end_color="FDEBD0", fill_type="solid")
light_purple = PatternFill(start_color="E8DAEF", end_color="E8DAEF", fill_type="solid")

white_font = Font(bold=True, size=11, color="FFFFFF")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)

def style_header_row(ws, row, cols, fill=dark_fill):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = fill
        cell.alignment = center
        cell.border = thin_border

def style_data_cell(ws, row, col, fill=None):
    cell = ws.cell(row=row, column=col)
    cell.font = normal_font
    cell.alignment = left_wrap
    cell.border = thin_border
    if fill:
        cell.fill = fill

def add_title(ws, row, text, cols, fill=dark_fill):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1)
    cell.value = text
    cell.font = title_font
    cell.fill = fill
    cell.alignment = center
    cell.border = thin_border
    for c in range(2, cols + 1):
        ws.cell(row=row, column=c).border = thin_border

# ========================================
# SHEET 1: Core Risk Rules
# ========================================
ws1 = wb.active
ws1.title = "Core Risk Rules"
ws1.sheet_properties.tabColor = "C0392B"

add_title(ws1, 1, "CORE RISK MANAGEMENT RULES", 4, red_fill)

headers = ["Rule #", "Rule", "Details", "Example (₹1,00,000 Account)"]
ws1.append(headers)
style_header_row(ws1, 2, 4)

rules = [
    ["1", "Max 2% risk per trade", "Never risk more than 2% of account on a single trade", "Max ₹2,000 per trade"],
    ["2", "Max 6% total exposure", "Maximum 3 positions open simultaneously", "3 trades × ₹2,000 = ₹6,000 max exposure"],
    ["3", "Daily loss limit: 4%", "If down 4% for the day, STOP trading", "Down ₹4,000 today → Stop, come back tomorrow"],
    ["4", "Weekly loss limit: 10%", "If down 10% for the week, STOP and review", "Down ₹10,000 this week → Stop, review strategy"],
    ["5", "Min Risk:Reward = 1:2", "Target must be 2x stop loss distance", "Stop 30 pips → Target 60+ pips"],
    ["6", "Max 3 trades per day", "Quality over quantity, prevents overtrading", "Even if 5 signals appear, only take best 3"],
    ["7", "Stop after 2 consecutive losses", "Prevents emotional revenge trading", "Loss + Loss → Stop for the day"],
]

for i, row in enumerate(rules):
    ws1.append(row)
    r = i + 3
    fill = light_red if i % 2 == 0 else None
    for c in range(1, 5):
        style_data_cell(ws1, r, c, fill)

ws1.column_dimensions['A'].width = 10
ws1.column_dimensions['B'].width = 30
ws1.column_dimensions['C'].width = 45
ws1.column_dimensions['D'].width = 45

# ========================================
# SHEET 2: Position Sizing Calculator
# ========================================
ws2 = wb.create_sheet("Position Sizing")
ws2.sheet_properties.tabColor = "2980B9"

add_title(ws2, 1, "POSITION SIZING CALCULATOR", 4, blue_fill)

headers = ["Step", "Formula", "Continuation Trade", "Reversal Trade"]
ws2.append(headers)
style_header_row(ws2, 2, 4)

sizing = [
    ["Step 1: Risk %", "Based on trade type", "1.5% (75% of capital)", "0.5% (25% of capital)"],
    ["Step 2: Risk Amount", "Account × Risk %", "₹1,00,000 × 1.5% = ₹1,500", "₹1,00,000 × 0.5% = ₹500"],
    ["Step 3: Stop Distance", "Entry - Stop Loss", "Entry ₹73.20, Stop ₹73.50 = 30 pips", "Entry ₹48,850, Stop ₹48,300 = 550 pips"],
    ["Step 4: Position Size", "Risk Amount ÷ Stop Distance", "₹1,500 ÷ 30 = ₹50/pip", "₹500 ÷ 550 = ₹0.91/pip"],
    ["Step 5: Lot Size", "Convert to lots (Forex)", "₹50/pip = 0.5 standard lots", "₹0.91/pip = 0.01 standard lots"],
]

for i, row in enumerate(sizing):
    ws2.append(row)
    r = i + 3
    fill = light_blue if i % 2 == 0 else None
    for c in range(1, 5):
        style_data_cell(ws2, r, c, fill)

# Capital Allocation section
ws2.append([])
add_title(ws2, 9, "CAPITAL ALLOCATION", 4, blue_fill)

headers2 = ["Trade Type", "Capital %", "Risk Per Trade", "Win Rate"]
r = 10
for c, h in enumerate(headers2, 1):
    ws2.cell(row=r, column=c, value=h)
style_header_row(ws2, r, 4)

alloc = [
    ["Continuation", "75% (₹75,000)", "1.5%", "60-70%"],
    ["Reversal", "25% (₹25,000)", "0.5%", "40-50%"],
]
for i, row in enumerate(alloc):
    ws2.append(row)
    r2 = 11 + i
    fill = light_green if i == 0 else light_orange
    for c in range(1, 5):
        style_data_cell(ws2, r2, c, fill)

ws2.column_dimensions['A'].width = 25
ws2.column_dimensions['B'].width = 30
ws2.column_dimensions['C'].width = 35
ws2.column_dimensions['D'].width = 35

# ========================================
# SHEET 3: Trade Management
# ========================================
ws3 = wb.create_sheet("Trade Management")
ws3.sheet_properties.tabColor = "27AE60"

add_title(ws3, 1, "TRADE MANAGEMENT RULES", 4, green_fill)

headers = ["Rule", "When to Apply", "Action", "Benefit"]
ws3.append(headers)
style_header_row(ws3, 2, 4)

mgmt = [
    ["1. Move to Breakeven", "Price reaches 1:1 R:R", "Move stop loss to entry price", "Risk-free trade"],
    ["2. Partial Profit (50%)", "Price reaches Target 1 (1:2 R:R)", "Close 50% of position", "Lock in profit, reduce stress"],
    ["3. Trail Stop", "After taking partial profit", "Trail stop by structure (swing highs/lows)", "Catch extended moves"],
    ["4. Max 3 Trades/Day", "Always", "Only take best 3 setups per day", "Prevents overtrading"],
    ["5. Stop After 2 Losses", "2 consecutive losing trades", "Stop trading for the day", "Prevents emotional trading"],
]

for i, row in enumerate(mgmt):
    ws3.append(row)
    r = i + 3
    fill = light_green if i % 2 == 0 else None
    for c in range(1, 5):
        style_data_cell(ws3, r, c, fill)

# Example section
ws3.append([])
add_title(ws3, 9, "EXAMPLE: SELL TRADE MANAGEMENT (From Your Chart)", 4, green_fill)

headers2 = ["Stage", "Price Level", "Action", "P&L"]
r = 10
for c, h in enumerate(headers2, 1):
    ws3.cell(row=r, column=c, value=h)
style_header_row(ws3, r, 4)

example = [
    ["Entry", "₹73.20", "SELL signal at Bearish OB", "₹0"],
    ["Stop Loss", "₹73.50", "30 pips above OB zone", "-₹1,500 (if hit)"],
    ["1:1 Reached", "₹72.90", "Move stop to ₹73.20 (breakeven)", "Risk-free now"],
    ["Target 1 (1:2)", "₹72.60", "Close 50% position", "+₹1,500 locked"],
    ["Trailing", "₹71.50", "Trail stop to ₹72.20", "Running profit"],
    ["Target 2 (1:5)", "₹71.70", "Close remaining 50%", "+₹3,750 total"],
]

for i, row in enumerate(example):
    ws3.append(row)
    r2 = 11 + i
    fill = light_green if i % 2 == 0 else light_yellow
    for c in range(1, 5):
        style_data_cell(ws3, r2, c, fill)

ws3.column_dimensions['A'].width = 25
ws3.column_dimensions['B'].width = 30
ws3.column_dimensions['C'].width = 45
ws3.column_dimensions['D'].width = 30

# ========================================
# SHEET 4: Drawdown Recovery
# ========================================
ws4 = wb.create_sheet("Drawdown Recovery")
ws4.sheet_properties.tabColor = "E67E22"

add_title(ws4, 1, "DRAWDOWN RECOVERY TABLE", 4, orange_fill)

headers = ["Loss %", "Amount Lost (₹1,00,000)", "Gain Needed to Recover", "Difficulty"]
ws4.append(headers)
style_header_row(ws4, 2, 4)

drawdown = [
    ["5%", "₹5,000", "5.3%", "Easy - Normal trading recovers this"],
    ["10%", "₹10,000", "11.1%", "Moderate - Reduce size, trade carefully"],
    ["15%", "₹15,000", "17.6%", "Hard - Stop trading, review strategy"],
    ["20%", "₹20,000", "25%", "Very Hard - Take 1 week break minimum"],
    ["30%", "₹30,000", "42.9%", "Critical - Backtest on demo first"],
    ["50%", "₹50,000", "100%", "Devastating - Major strategy overhaul needed"],
]

for i, row in enumerate(drawdown):
    ws4.append(row)
    r = i + 3
    if i <= 1:
        fill = light_green
    elif i <= 3:
        fill = light_orange
    else:
        fill = light_red
    for c in range(1, 5):
        style_data_cell(ws4, r, c, fill)

# Recovery Plan
ws4.append([])
add_title(ws4, 10, "RECOVERY PLAN", 4, orange_fill)

headers2 = ["Drawdown Level", "Action", "Position Size", "Signal Quality"]
r = 11
for c, h in enumerate(headers2, 1):
    ws4.cell(row=r, column=c, value=h)
style_header_row(ws4, r, 4)

recovery = [
    ["Down 5%", "Continue trading, review last 10 trades", "Normal (1.5%)", "High Only"],
    ["Down 10%", "STOP immediately, review last 20 trades", "Reduce by 50% (0.75%)", "Perfect Only"],
    ["Down 20%", "STOP for 1 week, complete strategy review", "Reduce by 75% (0.37%)", "Perfect Only"],
    ["Down 30%+", "STOP for 2 weeks, demo trade, get mentor", "Minimum size only", "Perfect Only + Mentor review"],
]

for i, row in enumerate(recovery):
    ws4.append(row)
    r2 = 12 + i
    if i == 0:
        fill = light_green
    elif i == 1:
        fill = light_yellow
    elif i == 2:
        fill = light_orange
    else:
        fill = light_red
    for c in range(1, 5):
        style_data_cell(ws4, r2, c, fill)

ws4.column_dimensions['A'].width = 20
ws4.column_dimensions['B'].width = 40
ws4.column_dimensions['C'].width = 35
ws4.column_dimensions['D'].width = 40

# ========================================
# SHEET 5: Signal Guide
# ========================================
ws5 = wb.create_sheet("Signal Guide")
ws5.sheet_properties.tabColor = "8E44AD"

add_title(ws5, 1, "SIGNAL TYPES & ZONE GUIDE", 5, purple_fill)

headers = ["Signal", "Label Color", "Meaning", "Action", "Risk Level"]
ws5.append(headers)
style_header_row(ws5, 2, 5)

signals = [
    ["💎 BUY", "Blue (Up)", "Bullish continuation from OB/FVG", "Enter LONG", "Low-Medium"],
    ["💎 SELL", "Orange (Down)", "Bearish continuation from OB/FVG", "Enter SHORT", "Low-Medium"],
    ["🚀 BUY REVERSAL", "Lime (Up)", "Bullish reversal (LQ grab + BOS)", "Enter LONG", "Medium-High"],
    ["🔻 SELL REVERSAL", "Fuchsia (Down)", "Bearish reversal (LQ grab + BOS)", "Enter SHORT", "Medium-High"],
]

for i, row in enumerate(signals):
    ws5.append(row)
    r = i + 3
    fill = light_blue if i < 2 else light_purple
    for c in range(1, 6):
        style_data_cell(ws5, r, c, fill)

# Zone colors
ws5.append([])
add_title(ws5, 8, "ZONE COLOR GUIDE", 5, purple_fill)

headers2 = ["Color", "Type", "Meaning", "Appears When", "Invalidated When"]
r = 9
for c, h in enumerate(headers2, 1):
    ws5.cell(row=r, column=c, value=h)
style_header_row(ws5, r, 5)

zones = [
    ["Green Solid Box", "Bullish Order Block", "Support in uptrend", "4+ bullish candles + HTF bullish", "Price closes below zone"],
    ["Red Solid Box", "Bearish Order Block", "Resistance in downtrend", "4+ bearish candles + HTF bearish", "Price closes above zone"],
    ["Blue Dashed Box", "Bullish FVG", "Support (price imbalance)", "Gap > 0.2% + HTF bullish", "Price fills the gap (low touches)"],
    ["Purple Dashed Box", "Bearish FVG", "Resistance (price imbalance)", "Gap > 0.2% + HTF bearish", "Price fills the gap (high touches)"],
]

for i, row in enumerate(zones):
    ws5.append(row)
    r2 = 10 + i
    fills = [light_green, light_red, light_blue, light_purple]
    for c in range(1, 6):
        style_data_cell(ws5, r2, c, fills[i])

# Confirmation checklist
ws5.append([])
add_title(ws5, 15, "ENTRY CONFIRMATION CHECKLIST", 5, purple_fill)

headers3 = ["#", "Check", "BUY Signal", "SELL Signal", "Status"]
r = 16
for c, h in enumerate(headers3, 1):
    ws5.cell(row=r, column=c, value=h)
style_header_row(ws5, r, 5)

checklist = [
    ["1", "HTF Bias matches direction", "Must be 🟢 BULLISH", "Must be 🔴 BEARISH", "☐"],
    ["2", "Signal quality High/Perfect", "High Only or Perfect Only", "High Only or Perfect Only", "☐"],
    ["3", "Clear zone visible", "Green OB or Blue FVG", "Red OB or Purple FVG", "☐"],
    ["4", "Confirmation candle", "Bullish engulfing or 70%+ body", "Bearish engulfing or 70%+ body", "☐"],
    ["5", "Risk:Reward min 1:2", "Target ≥ 2x stop distance", "Target ≥ 2x stop distance", "☐"],
    ["6", "No major news in 30 min", "Check economic calendar", "Check economic calendar", "☐"],
    ["7", "Good session timing", "London/NY overlap best", "London/NY overlap best", "☐"],
]

for i, row in enumerate(checklist):
    ws5.append(row)
    r2 = 17 + i
    fill = light_yellow if i % 2 == 0 else None
    for c in range(1, 6):
        style_data_cell(ws5, r2, c, fill)

ws5.column_dimensions['A'].width = 20
ws5.column_dimensions['B'].width = 25
ws5.column_dimensions['C'].width = 35
ws5.column_dimensions['D'].width = 35
ws5.column_dimensions['E'].width = 25

# ========================================
# SHEET 6: Trade Journal Template
# ========================================
ws6 = wb.create_sheet("Trade Journal")
ws6.sheet_properties.tabColor = "2C3E50"

add_title(ws6, 1, "TRADE JOURNAL", 8, dark_fill)

headers = ["Date", "Pair/Symbol", "Signal Type", "HTF Bias", "Entry", "Stop Loss", "Target", "Result"]
ws6.append(headers)
style_header_row(ws6, 2, 8)

# Add empty rows with formatting for user to fill
for i in range(20):
    row_num = i + 3
    for c in range(1, 9):
        cell = ws6.cell(row=row_num, column=c, value="")
        cell.border = thin_border
        cell.alignment = center
        if i % 2 == 0:
            cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")

# Add a details section below
ws6.append([])
add_title(ws6, 24, "DETAILED TRADE LOG", 8, dark_fill)

headers2 = ["Date", "Entry Reason", "Zone Type (OB/FVG)", "R:R Ratio", "Position Size", "P&L (₹)", "Emotions", "Lessons Learned"]
r = 25
for c, h in enumerate(headers2, 1):
    ws6.cell(row=r, column=c, value=h)
style_header_row(ws6, r, 8)

for i in range(20):
    row_num = 26 + i
    for c in range(1, 9):
        cell = ws6.cell(row=row_num, column=c, value="")
        cell.border = thin_border
        cell.alignment = left_wrap
        if i % 2 == 0:
            cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")

ws6.column_dimensions['A'].width = 15
ws6.column_dimensions['B'].width = 20
ws6.column_dimensions['C'].width = 20
ws6.column_dimensions['D'].width = 15
ws6.column_dimensions['E'].width = 15
ws6.column_dimensions['F'].width = 15
ws6.column_dimensions['G'].width = 15
ws6.column_dimensions['H'].width = 15

# ========================================
# SHEET 7: Quick Reference Card
# ========================================
ws7 = wb.create_sheet("Quick Reference")
ws7.sheet_properties.tabColor = "F39C12"

add_title(ws7, 1, "ICT SYSTEM - QUICK REFERENCE CARD", 3, orange_fill)

ws7.append([])

sections = [
    ("STEP 1: CHECK HTF BIAS", blue_fill, [
        ["🟢 BULLISH", "Price above HTF MA", "Only BUY setups"],
        ["🔴 BEARISH", "Price below HTF MA", "Only SELL setups"],
    ]),
    ("STEP 2: WAIT FOR SIGNAL", green_fill, [
        ["💎 BUY/SELL", "Continuation (OB/FVG bounce)", "Safer, 60-70% win rate"],
        ["🚀 BUY / 🔻 SELL REVERSAL", "Trend change (LQ + BOS)", "Riskier, 40-50% win rate"],
    ]),
    ("STEP 3: VERIFY SETUP", purple_fill, [
        ["HTF matches direction", "Zone visible (OB/FVG)", "Confirmation candle present"],
        ["R:R minimum 1:2", "No major news in 30 min", "Good session (London/NY)"],
    ]),
    ("STEP 4: EXECUTE TRADE", red_fill, [
        ["Entry", "Signal candle close", "Or next candle open"],
        ["Stop Loss", "Below zone (continuation)", "Below wick (reversal)"],
        ["Target 1", "Next structure (1:2)", "Take 50% profit"],
        ["Target 2", "Major level (1:3+)", "Trail remaining"],
    ]),
    ("STEP 5: MANAGE TRADE", dark_fill, [
        ["At 1:1", "Move stop to breakeven", "Risk-free trade"],
        ["At 1:2 (Target 1)", "Close 50% position", "Lock in profit"],
        ["After Target 1", "Trail stop with structure", "Catch extended moves"],
    ]),
]

current_row = 3
for title, fill, rows in sections:
    add_title(ws7, current_row, title, 3, fill)
    current_row += 1
    for row in rows:
        for c, val in enumerate(row, 1):
            cell = ws7.cell(row=current_row, column=c, value=val)
            cell.font = normal_font
            cell.alignment = center
            cell.border = thin_border
        current_row += 1
    current_row += 1

ws7.column_dimensions['A'].width = 30
ws7.column_dimensions['B'].width = 35
ws7.column_dimensions['C'].width = 35

# ========================================
# SAVE
# ========================================
filepath = r"C:\Users\Client\Desktop\ICT_Risk_Management_Guide.xlsx"
wb.save(filepath)
print(f"File saved: {filepath}")
